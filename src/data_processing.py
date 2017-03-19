import numpy as np
import openpyxl as xl
import os
import sys
import csv

def sort_data():

    root_path = '../Experimental_Data/Updated_Tests_Small_Motor/' #List root directory of data
    # root_path = '../Experimental_Data/Final/Small_motor_test/'
    prop_types = os.listdir(root_path)  #List all folders in directory
    dataBook = xl.Workbook()
    out_name = 'compiled_data.xlsx'



    i = 0
    for prop in prop_types:
        i += 1
        if '.' in prop:
            continue

        file_list = os.listdir(root_path + prop + '/')  #list files in each prop folder
        # try:
        #     file_list.remove('.DS_Store') #Remove .DS_Store from directory
        # except ValueError:
        #     pass

        #Set flag names (as seen in raw data), dictinary keys (user defined), and initialize data dictionary
        flags = ['Motor Electrical Speed (RPM)',
        'Current (A)', 
        'Thrust (gf)', 
        'Time (s)']

        keys = ['RPM', 'I', 'T', 't']
        raw_dict = dict(zip(flags, [ [], [], [], [] ]))

        data_keys = ['n', 'CT', 'CQ']
        data_dict = dict(zip(data_keys, [ [], [], [] ]))   

        for name in file_list:

            file_name = root_path + prop + '/' + name #Concatenate file name 

            if '.xlsx' in name:
                
                wb = xl.load_workbook(file_name)        #Load excel workbook

                sheet_name =  wb.sheetnames[0]      #Get 1st sheet name
                ws = wb.get_sheet_by_name(name = sheet_name)    #Load in 1st excel sheet
                
                titles = list(ws.rows)[0]   #Get a list of all of the titles of columns in the raw data sheet

                for i in flags:
                
                    for j, x in enumerate(titles):
                        if titles[j].value == i:
                            col = j #Match column index to flags

                    for row in ws.rows:
                        if isinstance(row[0].value, basestring):
                            continue
                        elif row[0].value is None:
                            break

                        raw_dict[i].append(row[col].value) #Add raw data to arrays in data dictionary

            elif '.csv' in name:
                
                with open(file_name, 'rb') as f:
                    reader = csv.reader(f)
                    rows = []
                    for r in reader:
                        rows.append(r)

                titles = rows[0]

                for i in flags:
                    
                    for j, x in enumerate(titles):
                        if titles[j] == i:
                            col = j #Match column index to flags

                    row_floats = []
                    for row in rows:
                        try:
                            var = float(row[0])
                        except ValueError:
                            continue
                   
                        if bool(row[0]):

                            for j in titles:
                                if i in j:

                                    data_col = titles.index(j)
                                    row_floats.append(float(row[data_col]))
                            
                        else:
                            break

                        raw_dict[i] = row_floats #Add raw data to arrays in data dictionary

            else:
                continue

            name, ext = name.split('.') #Remove extension from name
            motor_name, prop_string, kv_string, volt_string =  name.split('_')  #Parse name into motor name, prop_string, kv_string, and volt_string

            prop_diameter = float(prop_string[0:2])*.1  #Use prop string to express diameter
            prop_pitch = float(prop_string[2:])*.1      #Use prop string to express pitch
            kv = float(kv_string)                       #Use kv string to express kv
            kt = 1352.4/kv                              #Use kv to compute kv by definition
            V = float(volt_string)*.1                   #Use volt string to express voltage


            for i, x in enumerate(keys):
                raw_dict[keys[i]] = raw_dict.pop(flags[i])    #Change keys in dictinary from raw data flags to user defined keys (optional)

            segment = [i for i,x in enumerate(raw_dict['t']) if x > 10 and x < 30] #Select segment of data in middle of test to reduce transients

            for i, x in raw_dict.iteritems():
                raw_dict[i] = raw_dict[i][segment[0]:segment[-1]+1]   #Slice the desire segment of data from all data arrays

            # raw_dict['T'] = [x*101.971621298 for x in raw_dict['T']]  #Convert thrust from [N] to [g]

            rho = .002377   #Define desnsity of air in slug/ft**3

            n_array = [x/60. for x in raw_dict['RPM']]  #Make array of rev/s

            #Compute thrust coefficient CT = T/(rho*(n**2)*(D**4))
            CT_array = [raw_dict['T'][i]/(rho*((prop_diameter/12.)**4.0)*(n_array[i]**2.)) for i,x in enumerate(raw_dict['T'])]

            tau_array = [kt*x for x in raw_dict['I']] #Compute torque based on raw data and motor kt [oz-in]

            #Compute torque coefficient CQ = tau/(rho*(n**2)*(D**5))
            CQ_array = [tau_array[i]/(rho*((prop_diameter/12.)**5.0)*(n_array[i]**2.)) for i,x in enumerate(tau_array)]

            #Append data to arrays in data_dict
            for i,x in enumerate(n_array):
                data_dict['n'].append(n_array[i])
                data_dict['CT'].append(CT_array[i])
                data_dict['CQ'].append(CQ_array[i])

            for i, x in enumerate(keys):
                raw_dict[flags[i]] = raw_dict.pop(keys[i])    #Change keys in dictinary back to raw data flags (optional)
                raw_dict[flags[i]] = []     #Reset raw dict


        #Write data to compiled workbook
        ws_new = dataBook.create_sheet(title = prop_string)

        #Remove dummy sheet
        try:
            std = dataBook.get_sheet_by_name('Sheet')   
            dataBook.remove_sheet(std)
        except KeyError:
            pass

        #Write headers
        for i,x in enumerate(data_keys):
            ws_new.cell(row = 1, column = i+1, value = x)

            myData = data_dict[x]
            for j, y in enumerate(myData):
                ws_new.cell(row = j+2, column = i+1, value = y)


        dataBook.save(root_path + out_name)

def write_data_textfiles():
    file_name = '../Experimental_Data/Updated_Tests_Small_Motor/compiled_data.xlsx'
    out_folder = '../prop_data/'

    wb = xl.load_workbook(file_name)
    sheet_list = wb.get_sheet_names()

    flags = ['n', 'CT', 'CQ']

    for sheet in sheet_list:

        ws = wb.get_sheet_by_name(sheet)
        titles = list(ws.rows)[0]
        ind_list = []

        for title in titles:
            if title.value in flags:
                ind_list.append(titles.index(title))

        write_data = []
        for row in ws.rows:
            if isinstance(row[0].value, basestring):
                continue
            elif row[0].value is None:
                break

            dataObj = [row[i] for i in ind_list]

            write_row = [i.value for i in dataObj]
            write_data.append(write_row)

        outfile = out_folder + sheet + '.csv'
        
        with open(outfile, 'wb') as f:
            writer = csv.writer(f, delimiter = ',')
            writer.writerow(flags)
            for row in write_data:
                writer.writerow(row)
        

if __name__ == '__main__':

    # sort_data()
    write_data_textfiles()

